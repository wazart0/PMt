import sys
import os
from atlassian import Jira
from bidict import bidict

from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime




filename = sys.argv[1]
label = sys.argv[2]


jira_info = {
    'url': sys.argv[3],
    'username': sys.argv[4],
    'api_token': sys.argv[5],
}



path = os.path.dirname(os.path.realpath(__file__))



def datetime_to_date_converter(date):
    from datetime import datetime
    from dateutil import parser
    if isinstance(date, datetime):
        return date.strftime("%Y-%m-%d")
    return parser.parse(date).strftime("%Y-%m-%d")


def get_schema_from_sheet(sheet, type): # TODO schema validation
    schema = {}

    if type == 'FE':
        schema['offset_row'] = 1 # offset of WBS: TODO search for WBS
        schema['offset_col'] = 0 # offset of WBS: TODO search for WBS
        schema['columns'] = { # TODO make it dynamic
            'Category': 'A',
            'WBS': 'C',
            'Name': 'B',
            'Description': 'M',
            'Depends on': 'D',
            'Estimate [h]': 'E',
            'Jira ID': 'V',
            'Start time': 'S', # Planned and actual
            'Finish time': 'T', # Planned and actual
            # 'Duration [h]': 'L'
        }
        schema['data_offset_row'] = 3
    
    if type == 'BE' or type == 'QA':
        schema['offset_row'] = 1 # offset of WBS: TODO search for WBS
        schema['offset_col'] = 0 # offset of WBS: TODO search for WBS
        schema['columns'] = { # TODO make it dynamic
            'WBS': 'A',
            'Name': 'B',
            'Description': 'C',
            'Depends on': 'D',
            'Estimate [h]': 'H',
            'Jira ID': 'O',
            'Start time': 'J', # Planned and actual
            'Finish time': 'K', # Planned and actual
            'Duration [h]': 'L',
            'Category': 'S',
            'Sprint': 'Q'
        }
        schema['data_offset_row'] = 3

    return schema
    

def wbs_regex_check(string, separator = '.'):
    import re
    return True if re.match('\A([0-9]+\\' + separator + '?)+\Z', string) else False


def wbs_regex_task(string):
    import re
    return True if re.match('\A[0-9]+\Z', string) else False


def check_wbs_level(wbs, separator = '.'):
    return len(wbs.split(separator))


def get_wbs_ancestor(wbs, separator = '.'):
    if len(wbs.split(wbs)) == 1:
        return wbs.split(wbs)[0]
    else:
        return '.'.join(wbs.split('.')[0:-1])


def check_if_successor(wbs, successor):
    wbs_split = wbs.split('.')
    successor_split = successor.split('.')
    if len(wbs_split) >= len(successor_split):
        return False
    for i in range(len(wbs_split)):
        if wbs_split[i] != successor_split[i]:
            return False
    return True


def get_wbs_lowest_level_task_number(wbs, separator = '.'):
    return wbs.split(separator)[-1]


def expand_wbs_range(start, finish, separator = '.'):
    if check_wbs_level(start, separator) != check_wbs_level(finish, separator):
        print('ERROR: WBS range level differs: ', start, ' - ', finish)
        return []
    if get_wbs_ancestor(start, separator) != get_wbs_ancestor(finish, separator):
        print('ERROR: WBS range ancestor differs: ', start, ' - ', finish)
        return []
    lowest_level_task = list(range(int(get_wbs_lowest_level_task_number(start, separator)), 1 + int(get_wbs_lowest_level_task_number(finish, separator))))
    wbs_tasks = []
    for i in lowest_level_task:
        if get_wbs_ancestor(start, separator) == '':
            wbs_tasks.append(str(i))
        else:
            wbs_tasks.append(get_wbs_ancestor(start, separator) + '.' + str(i))
    return wbs_tasks
    

def parse_dependencies(dependency_str, wbs_separator = '.'):
    wbs_tasks = []
    for i in str(dependency_str).split(','):
        if wbs_regex_check(i.strip(), wbs_separator):
            wbs_tasks.append(i)
        else:
            range = i.split('-')
            if len(range) != 2:
                print("ERROR: It is not a range: ", range)
                continue
            if wbs_regex_check(range[0].strip(), wbs_separator) == False or wbs_regex_check(range[1].strip(), wbs_separator) == False:
                print("ERROR: It is not a range: ", range)
                continue
            wbs_tasks = wbs_tasks + expand_wbs_range(range[0].strip(), range[1].strip(), wbs_separator)
    return wbs_tasks


def get_value_for_wbs(schema, sheets, wbs, column_name):
    for sheet in sheets:
        i = schema['data_offset_row']
        none_count = 0
        while none_count < 50:
            if sheet[schema['columns']['WBS'] + str(i)].value is not None:
                if wbs == str(sheet[schema['columns']['WBS'] + str(i)].value):
                    return str(sheet[schema['columns'][column_name] + str(i)].value)

                none_count = 0
            else:
                none_count = none_count + 1
            i = i + 1 
    return None


def get_jira_key_for_wbs(schema, sheets, wbs):
    return get_value_for_wbs(schema, sheets, wbs, 'Jira ID')


def get_start_finish_from_child(sheets, schema, wbs):
    from dateutil import parser
    min = None
    max = None
    for sheet in sheets:
        i = schema['data_offset_row']
        none_count = 0
        while none_count < 50:
            if sheet[schema['columns']['WBS'] + str(i)].value is not None:

                if check_if_successor(wbs, str(sheet[schema['columns']['WBS'] + str(i)].value)):

                    if min is None and sheet[schema['columns']['Start time'] + str(i)].value is not None:
                        if isinstance(sheet[schema['columns']['Start time'] + str(i)].value, datetime.datetime):
                            min = sheet[schema['columns']['Start time'] + str(i)].value
                        else:
                            min = parser.parse(sheet[schema['columns']['Start time'] + str(i)].value)

                    if max is None and sheet[schema['columns']['Finish time'] + str(i)].value is not None:
                        if isinstance(sheet[schema['columns']['Finish time'] + str(i)].value, datetime.datetime):
                            max = sheet[schema['columns']['Finish time'] + str(i)].value
                        else:
                            max = parser.parse(sheet[schema['columns']['Finish time'] + str(i)].value)

                    if sheet[schema['columns']['Start time'] + str(i)].value is not None:
                        if isinstance(sheet[schema['columns']['Start time'] + str(i)].value, datetime.datetime):
                            if sheet[schema['columns']['Start time'] + str(i)].value < min:
                                min = sheet[schema['columns']['Start time'] + str(i)].value
                        else:
                            if parser.parse(sheet[schema['columns']['Start time'] + str(i)].value) < min:
                                min = parser.parse(sheet[schema['columns']['Start time'] + str(i)].value)

                    if sheet[schema['columns']['Finish time'] + str(i)].value is not None: 
                        if isinstance(sheet[schema['columns']['Finish time'] + str(i)].value, datetime.datetime):
                            if sheet[schema['columns']['Finish time'] + str(i)].value > max:
                                max = sheet[schema['columns']['Finish time'] + str(i)].value
                        else:
                            if parser.parse(sheet[schema['columns']['Finish time'] + str(i)].value) > max:
                                max = parser.parse(sheet[schema['columns']['Finish time'] + str(i)].value)

                none_count = 0
            else:
                none_count = none_count + 1

            i = i + 1 
    return min, max


def get_min_from_start(sheets, schema):
    from dateutil import parser
    min = None
    for sheet in sheets:
        i = schema['data_offset_row']
        none_count = 0
        while none_count < 50:
            if sheet[schema['columns']['WBS'] + str(i)].value is not None:
                if min is None and sheet[schema['columns']['Start time'] + str(i)].value is not None:
                    if isinstance(sheet[schema['columns']['Start time'] + str(i)].value, datetime.datetime):
                        min = sheet[schema['columns']['Start time'] + str(i)].value
                    else:
                        min = parser.parse(sheet[schema['columns']['Start time'] + str(i)].value)

                if sheet[schema['columns']['Start time'] + str(i)].value is not None: 
                    if isinstance(sheet[schema['columns']['Start time'] + str(i)].value, datetime.datetime):
                        if sheet[schema['columns']['Start time'] + str(i)].value < min:
                            min = sheet[schema['columns']['Start time'] + str(i)].value
                    else:
                        if parser.parse(sheet[schema['columns']['Start time'] + str(i)].value) < min:
                            min = parser.parse(sheet[schema['columns']['Start time'] + str(i)].value)

                none_count = 0
            else:
                none_count = none_count + 1
            i = i + 1 
    return min





def schema_FE(sheet, schema, index, task_json):
    if sheet[schema['columns']['Start time'] + str(index)].value is not None:
        task_json['customfield_11701'] = datetime_to_date_converter(sheet[schema['columns']['Start time'] + str(index)].value)
    if sheet[schema['columns']['Finish time'] + str(index)].value is not None:
        task_json['customfield_11702'] = datetime_to_date_converter(sheet[schema['columns']['Finish time'] + str(index)].value)

    if sheet[schema['columns']['Start time'] + str(index)].value is not None:
        task_json['customfield_11724'] = datetime_to_date_converter(sheet[schema['columns']['Start time'] + str(index)].value)
    if sheet[schema['columns']['Finish time'] + str(index)].value is not None:
        task_json['customfield_11725'] = datetime_to_date_converter(sheet[schema['columns']['Finish time'] + str(index)].value)

    if sheet[schema['columns']['Estimate [h]'] + str(index)].value is not None:
        task_json['customfield_10004'] = float(sheet[schema['columns']['Estimate [h]'] + str(index)].value)/4
        
    return task_json


def create_or_update_issue(sheets, sheet, schema, index, project, type, label, schema_type, parent = None, jira = None):

    print("current sheet: ", sheet, "    current row: ", index)

    ### form basic fields for create or update
    task_json = {
        'summary': sheet[schema['columns']['Name'] + str(index)].value,
        'description': sheet[schema['columns']['Description'] + str(index)].value,
        'components': [{"name": schema_type}]
    }

    if schema_type == 'FE':
        task_json = schema_FE(sheet, schema, index, task_json)

    ## create issue
    if sheet[schema['columns']['Jira ID'] + str(index)].value is None: 

        task_json['project'] = {'key': project}
        task_json['issuetype'] = {'name': type}
        if parent is not None:
            task_json['parent'] = {'key': parent}

        if jira is not None:
            issue = jira.issue_create(task_json)
            # print(issue)
            sheet[schema['columns']['Jira ID'] + str(index)].value = issue['key']
            sheet[schema['columns']['Jira ID'] + str(index)].hyperlink = jira.url + '/browse/' + issue['key']
            sheet[schema['columns']['Jira ID'] + str(index)].font = Font(underline='single', color='0000FF')

    ## update basic fields     
    else:        
        if jira is not None:    
            jira.issue_update(sheet[schema['columns']['Jira ID'] + str(index)].value, task_json)
        # remove_all_links(jira_info, sheet[schema['columns']['Jira ID'] + str(index)].value)

    # print(task_json)

    ### section for edit issue

    ## add dependencies
    if sheet[schema['columns']['Depends on'] + str(index)].value is not None:
        if sheet[schema['columns']['Depends on'] + str(index)].value != '':
            # print(parse_dependencies(sheet[schema['columns']['Depends on'] + str(index)].value))
            for dependent_task_wbs in parse_dependencies(sheet[schema['columns']['Depends on'] + str(index)].value):
                dependent_jira_id = get_jira_key_for_wbs(schema, sheets, dependent_task_wbs.strip())
                if dependent_jira_id is not None:
                    # print(dependent_jira_id)
                    fields = {
                        "issuelinks": [
                            {
                                "add": {
                                    "type": {
                                        "name": "Gantt End to Start",
                                        "inward": "has to be done after",
                                        "outward": "has to be done before"
                                    },
                                    "inwardIssue": {
                                        "key": dependent_jira_id
                                    }
                                }
                            }
                        ]
                    }
                    # add dependency
                    if jira is not None:
                        jira.edit_issue(issue_id_or_key=sheet[schema['columns']['Jira ID'] + str(index)].value, fields=fields, notify_users=False)

    ## add labels
    labels = [{"add": label}]
    if sheet[schema['columns']['Category'] + str(index)].value is not None:
        labels = labels + [{"add": (str(sheet[schema['columns']['Category'] + str(index)].value)).strip().replace(' ', '_')}]

    if jira is not None:
        jira.edit_issue(issue_id_or_key=sheet[schema['columns']['Jira ID'] + str(index)].value, fields={"labels": labels}, notify_users=False)



def create_or_update_issues(jira, sheets, schema, project, label, schema_type):
    for sheet in sheets:
        print("current sheet: ", sheet)
        i = schema['data_offset_row']
        none_count = 0
        while none_count < 50:
            if sheet[schema['columns']['WBS'] + str(i)].value is not None:
                wbs = str(sheet[schema['columns']['WBS'] + str(i)].value)
                if not wbs_regex_check(wbs):
                    print('Problem in WBS number structure: ', wbs, ' in sheet: ', sheet)
                    i = i + 1
                    continue

                if sheet[schema['columns']['Name'] + str(i)].value is None or sheet[schema['columns']['Name'] + str(i)].value == '':
                    print('Task do not have any name, wbs: ', wbs, ' in sheet: ', sheet)
                    i = i + 1
                    continue

                create_or_update_issue(jira=jira, sheets=sheets, sheet=sheet, schema=schema, index=i, project=project, type='Task', label=label, schema_type=schema_type)

                none_count = 0
            else:
                none_count = none_count + 1
            i = i + 1




if __name__ == "__main__":

    jira = Jira(
            url=jira_info['url'],
            username=jira_info['username'],
            password=jira_info['api_token'],
            cloud=True)

    filename = 'P2 Unit Test Estimations.xlsx'
    label = 'P2v3frontend'
    schema_type = 'FE'


    exclude_tab = ["PMSummary"]

    workbook = load_workbook(filename=path+'/'+filename, data_only=True)

    sheets_to_analyze = []

    for sheet in workbook.sheetnames:
        if sheet not in exclude_tab:
            sheets_to_analyze.append(workbook[sheet])

    schema = get_schema_from_sheet(sheets_to_analyze, schema_type)
    create_or_update_issues(jira=jira, sheets=sheets_to_analyze, schema=schema, project='PP', label=label, schema_type=schema_type)
    
    workbook.save(filename=filename)
    workbook.close()